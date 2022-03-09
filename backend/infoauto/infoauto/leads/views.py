import datetime
import json
from tempfile import NamedTemporaryFile

from django.apps import apps
from django.core.exceptions import ObjectDoesNotExist
from django.core.files.storage import FileSystemStorage
from django.db import transaction
from django.db.models import Q, ForeignKey, F, Count, Sum, DateTimeField, IntegerField, Avg, DurationField
from django.db.models.expressions import RawSQL, Value, ExpressionWrapper, Case, When
from django.db.models.functions import Cast, TruncSecond, TruncHour, Extract, Substr, Concat
from django.http import HttpResponse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.translation import ugettext_lazy as _
from django_filters.rest_framework import DjangoFilterBackend
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl import Workbook
from rest_condition import Or, permissions
from rest_framework import mixins, status, serializers, authentication
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.generics import get_object_or_404
from rest_framework.permissions import DjangoModelPermissions, IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import GenericViewSet as GenericViewSetAux

from infoauto.common.permissions import IsAdminUser
from infoauto.leads.filters import LeadFilter, LeadColsFilter, LeadActionFilter, LeadManagementFilter, \
    LeadCalendarFilter, LeadMasterFilter, TaskFilter
from infoauto.leads.models import LEAD_RESULT, LEAD_STATUS_NEW
from infoauto.leads.models.leads import LeadAction, LeadCalendar, INTEGRATED, NOT_HBS_API_KEY, LeadWhatsAppMessage
from infoauto.leads.serializers.core import LeadTaskCreateSerializer, CommonTaskSerializer
from infoauto.leads.serializers.lead_actions import LeadActionSerializer
from infoauto.leads.serializers.leads import ExcelLeadSerializer, LeadCalendarSerializer, LeadMasterSerializer, \
    WhatsAppSerializer
from infoauto.vehicles.serializers import GasTypeSerializer
from .lead_master import LeadMaster, LeadMasterPersistente
from .models.lead_management import LeadManagement, EVENT_LEAD_REACTIVATED_KEY

from .permissions import IsConcessionaireAdmin
from .models import Lead, Task, LEAD_STATUS, Request, GasType, ACD, Origin
from infoauto.leads.serializers import LeadSerializer, TaskSerializer, TaskCreateSerializer, \
    LeadColumnSerializer, ACDSerializer, OriginSerializer, LeadStatusSerializer, \
    HistorySerializer

from .serializers.lead_importer_serializers import LeadImporterSerializer
from .utils import special_date_ranges, create_excel_template, add_value_to_excel, add_column_to_excel

from rest_framework.serializers import ValidationError
from hubspot3 import Hubspot3
from hubspot3.error import HubspotConflict, HubspotNotFound, HubspotBadRequest

import base64
import xlsxwriter
import json
import io
import xlrd


def model_clone(instance, avoid=None, set_none=None):
    """
    :param instance: model instance
    :param avoid: ForeignKey fields to set without clone (uses same instance)
    :param set_none: set attributes to None
    :return: instance
    """
    avoid = avoid or {}
    set_none = set_none or {}
    instance.id = None
    field_names = []
    for i in instance._meta.model._meta.fields:
        if isinstance(i, ForeignKey) and i.name not in set_none.keys():
            if (i.name not in avoid.keys()) or (i.name in avoid.keys() and avoid.get(i.name)):
                if (i.name not in set_none.keys()) or (i.name in set_none.keys() and set_none[i.name]):
                    field_names.append(i.name)
        elif i.name in set_none.keys() and not set_none[i.name]:
            setattr(instance, i.name, None)
    for i in field_names:
        if getattr(instance, i, None):
            attr_inst = model_clone(instance=getattr(instance, i, None), avoid=avoid.get(i, None),
                                    set_none=set_none.get(i, None))
            setattr(instance, i, attr_inst)
    instance.save()
    return instance


class GenericViewSet(GenericViewSetAux):

    def check_lead_status(self, request, lead):
        if not request.user.is_admin:
            if not request.user.userconcession_set.filter(concessionaire=lead.concessionaire).exists():
                raise ValidationError({"non_field_errors": [_("You are not member of this concession")]})
            elif lead.status in ['new', 'pending'] and not lead.user:
                self.permission_classes = [Or(IsAdminUser, IsConcessionaireAdmin)]
                self.check_permissions(request)


class LeadGenericViewSet(GenericViewSet):
    raw_sql = '''
        SELECT `la2`.`date`
        FROM `leads_lead` as `l`
        left JOIN (SELECT `la`.`lead_id` as `lead_id`, MAX(`la`.`id`) as `max_id`
                    from `leads_leadaction` `la`
                    GROUP BY `la`.`lead_id`) `la` on `la`.`lead_id` = `l`.`id`
        LEFT JOIN `leads_leadaction` `la2` on `la`.`max_id` = `la2`.`id`
        where `l`.`id` = `leads_lead`.`id`
        group by `l`.`id`
    '''

    def get_queryset(self, *args, **kwargs):
        if self.request.user.is_authenticated and not self.request.user.is_superuser:
            concessionaires = list(self.request.user.userconcession_set.filter(
                Q(is_concessionaire_admin=True)
            ).distinct().values_list('concessionaire', flat=True))
            self.queryset = self.queryset.filter(
                Q(concessionaire__id__in=concessionaires) | Q(user__id=self.request.user.id)
            ).distinct()
        elif not self.request.user.is_authenticated:
            self.queryset = self.queryset.none()
        self.queryset = self.queryset.annotate(
            last_lead_action_date=RawSQL(self.raw_sql, params=())
        ).order_by('last_lead_action_date')
        return super().get_queryset()


class LeadView(mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.CreateModelMixin, mixins.UpdateModelMixin,
               LeadGenericViewSet):
    serializer_class = LeadSerializer
    queryset = Lead.objects.all().order_by("id")
    permission_classes = [Or(IsAdminUser, DjangoModelPermissions, IsAuthenticated)]
    filter_backends = [SearchFilter, DjangoFilterBackend, OrderingFilter]
    filter_class = LeadFilter
    search_fields = ('status', 'vehicles__brand_model', 'client__name', 'client__surname', 'client__phone', 'client__desk_phone', 'request__task__description',
                     'request__task__realization_date', 'request__task__realization_date', 'tags__content', 'tags__id',
                     'client__business_name')
    ordering_fields = (
        'created',
        'finish_date',
        'modified',
        'client__name',
        'client__surname',
        'status',
    )

    # START - Excel export
    map_names = {
        'status': "Estado", "client_name": "Nombre del cliente", "client_phone": "Teléfono",
        "vehicle_brand_model": "Vehículo demandada modelo marca",
        "vehicle_price": "Vehículo demandado precio", "vehicle_km": "Vehículo demandado kilometro",
        "vehicle_year": "Vehículo demandado año", "vehicle_gas": "Vehículo demandado combustible",
        "notes": "Notas", "owned_by": "Asignado a", "score": "Probabilidad de compra",
        "result": "Resultado", "created": "Fecha de alta", "nsa": "NSA",
        "last_realization_task": "Fecha ultima tarea", "last_tracking_task": "Fecha ultimo seguimiento",
        "nameconcession": "Nombre de la Concesión", "origin": "Origen", "medio": "Medio", "province": "Provincia",
        "historic": "Histórico", 'cita': "Cita", "tasacion": "Tasación", "presupuesto": "Presupuesto",
        "financiacion": "Financiación", "garantia": "Garantía", "info_vehiculo": "Información de Vehículo",
        "respuesta_mail": "Respuesta con información mediante mail", "otros": "Otros"
    }
    map_values = {**dict(LEAD_STATUS), **dict(LEAD_RESULT)}

    # END - Excel export

    def get_serializer_class(self):
        self.serializer_class = LeadSerializer
        if self.action == 'change_status':
            self.serializer_class = LeadStatusSerializer
        elif self.action in ['excel']:
            self.serializer_class = ExcelLeadSerializer
        return super().get_serializer_class()

    def list(self, request, *args, **kwargs):
        if not request.GET.get('with_concession', False):
            self.queryset = self.get_queryset().filter(user=request.user)
        return super().list(request, *args, **kwargs)

    @swagger_auto_schema(responses={status.HTTP_200_OK: json.dumps(['string'])})
    @action(methods=['GET'], detail=False)
    def lead_status(self, request, *args, **kwargs):
        return Response(dict(LEAD_STATUS).keys(), status=status.HTTP_200_OK)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.check_lead_status(request, instance)
        return super().update(request, *args, **kwargs)

    @action(methods=['PATCH'], detail=True)
    def change_status(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)

    @action(methods=['POST'], detail=True)
    def reactivate(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.reactivate()
        return Response(LeadSerializer(instance, context={'request': request}).data)

    @action(methods=['POST'], detail=True)
    def email(self, request, *args, **kwargs):
        instance = self.get_object()
        LeadManagement.objects.create(
            lead_id=instance.id,
            event=EVENT_LEAD_REACTIVATED_KEY,
            user=instance.user,
            status=instance.status,
            message="Email enviado"
        )
        return Response(status=status.HTTP_201_CREATED)

    @action(methods=["GET"], detail=False)
    def calendar(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())[:100]
        serializer = self.get_serializer(queryset, many=True)
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(methods=["GET"], detail=False)
    def excel(self, request, *args, **kwargs):
        self.serializer_class = LeadMasterSerializer

        if not request.GET.get('with_concession', False):
            self.queryset = self.get_queryset().filter(user=request.user)

        self.queryset = self.filter_queryset(self.queryset)

        lead_ids = [lead.id for lead in self.queryset]
        self.queryset = LeadMasterPersistente.objects.filter(id__in=lead_ids).order_by("id")
        columns = []
        colum_names = []
        fields = LeadMasterPersistente._meta.fields
        excluded_clolums = ['concessionaire', 'created']

        for field in fields:
            if field.name not in excluded_clolums:
                columns.append(field.name)
                colum_names.append(field.db_column if field.db_column else field.name)

        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output)
        output_ws = wb.add_worksheet('Leads Export')
        queryset = self.queryset

        for y, column_name in enumerate(colum_names):
            add_column_to_excel(column_name, output_ws, wb, y)

        for x, obj in enumerate(queryset, start=1):
            for y, column in enumerate(columns):
                value = getattr(obj, column) if getattr(obj, column) else ""
                add_value_to_excel(str(value), output_ws, x, y)

        wb.close()
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = 'attachment; filename=%s' % "lead_export"
        return response

    @action(methods=["POST"], detail=True)
    def update_hubspot(self, request, *args, **kwargs):
        instance = self.get_object()

        if request.user.is_authenticated:
            vendedor_digital = instance.user.first_name  # request.data.get("vendedor_digital")
        else:
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": _("Missing Digital seller")})

        if not vendedor_digital:
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": _("Missing Digital seller")})

        if not instance.client:
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": _("This lead has not client")})

        # HubSpot integration status
        if instance.hubspot_integration_status == INTEGRATED:
            return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": _("This lead is updated")})
        elif instance.hubspot_integration_status == NOT_HBS_API_KEY:
            return Response(status=status.HTTP_400_BAD_REQUEST,
                            data={"error": _("This concessionaire has not HubSpot API Key")})
        else:
            # Update hubspot deals process
            hubspot = self._get_hubspot_instance(instance)

            # Validations
            # 1.- Get hubspot property
            # 3.- Get property options
            # 4.- Validate if data value is in properties

            # Personal data
            if not instance.client.phone:
                return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": "Phone is required"})
            if not instance.client.get_full_name():
                return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": "Name is required"})
            if not instance.client.email:
                return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": "Email is required"})

            # Vendedor digital
            vendedor_digital_property = hubspot.properties.get("contacts", "vendedor_digital")
            vendedor_digital_options = [option.get("value") for option in vendedor_digital_property.get("options")]
            if vendedor_digital not in vendedor_digital_options:
                return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": _(
                    "Vendedor digital (" + vendedor_digital + ") is not in option list values: " + ", ".join(
                        vendedor_digital_options))})

            # VN/VO
            vn_vo = instance.vehicles.first().vehicle_type
            if vn_vo in ["km0", "seminew", "used"]:
                vn_vo = "VO"
            else:
                vn_vo = "VN"

            # Origen
            origen = instance.source.origin.name
            origen_property = hubspot.properties.get("contacts", "origenoculto")
            origen_property_values = [option.get("value") for option in origen_property.get("options")]

            matching = [s for s in origen_property_values if origen.upper() in s.upper()]

            if len(matching) == 1:
                origen = matching[0]
            elif len(matching) > 1:
                origen = [match for match in matching if match.upper() == "Llamada {} {}".format(origen, vn_vo).upper()]
                if len(origen) > 0:
                    origen = origen[0]

            # marcaoculto
            if not instance.vehicles.first().brand_model:
                return Response(status=status.HTTP_400_BAD_REQUEST,
                                data={"error": _("Vehicle brand product is required")})

            marcaoculto_property = hubspot.properties.get("contacts", "marcaoculto")
            marcaoculto_options = [option.get("value") for option in marcaoculto_property.get("options")]

            if instance.vehicles.first().brand_model.split("___")[0].upper() not in marcaoculto_options:
                return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": _(
                    "marcaoculto is not in option list values: " + ", ".join(marcaoculto_options))})

            # concesion
            concesi_n_property = hubspot.properties.get("contacts", "concesi_n")
            concesi_n_options = [option.get("value") for option in concesi_n_property.get("options")]
            if instance.concessionaire.name not in concesi_n_options:
                return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": _(
                    "Concesion is not in option list values: " + ", ".join(concesi_n_options))})

            # Static values
            LEGAL_BASE_DEFAULT = "Legitimate interest – prospect/lead"
            LEAD_STATUS_DEFAULT = "NEW"
            LIFE_CICLE_STATUS_DEFAULT = "lead"
            ORIGINAL_SOURCE_DEFAULT = "REFERRALS"

            if instance.status == LEAD_STATUS_NEW:
                llamada_atendida_ = "No"
            else:
                llamada_atendida_ = "Sí"

            contact_values = {
                "origenoculto": origen,
                "concesi_n": instance.concessionaire.name,
                "vn_vo": vn_vo,
                "mobilephone": instance.client.phone,
                "firstname": instance.client.get_full_name(),
                "email": instance.client.email,
                "message": ". ".join('{}. '.format(note.content) for note in instance.note.all()),
                "marcaoculto": instance.vehicles.first().brand_model.split("___")[0].upper(),
                "vendedor_digital": vendedor_digital,
                "llamada_atendida_": llamada_atendida_,
                "hs_legal_basis": LEGAL_BASE_DEFAULT,
                "hs_lead_status": LEAD_STATUS_DEFAULT,
                "lifecyclestage": LIFE_CICLE_STATUS_DEFAULT,
                "hs_analytics_source": ORIGINAL_SOURCE_DEFAULT,
            }

            # Contact creation to assign deal
            contact = None
            try:
                contact = hubspot.contacts.create(
                    data={"properties": [{"property": k, "value": v} for k, v in contact_values.items()]})
            except HubspotBadRequest as e:
                error_msg = json.loads(e.result.body).get('validationResults')
                error_txt = ""
                pattern = "(((value)+: ((\")?([a-zA-Z0-9\\ \.]+)(\")?\n)+))"
                import re

                for error in error_msg:
                    print(error)
                    error_txt += error.get("name") + str(_(". Values are: "))
                    parsed_message = re.findall(pattern, error.get('message'))
                    for p in parsed_message:
                        error_txt += p[5] + ", "
                return Response(status=status.HTTP_400_BAD_REQUEST,
                                data={"error": error_txt})
            except HubspotConflict:
                contact = hubspot.contacts.get_by_email(instance.client.email)

            # company = None
            # try:
            #     company = hubspot.companies.get(instance.concessionaire.hubspot_id)
            # except HubspotNotFound:
            #     print("Company not found")
            #     return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": _("Company not found")})

            # if contact and company:
            #     deal = hubspot.deals.create({
            #         "associations": {
            #             "associatedCompanyIds": [
            #                 company.get("companyId")
            #             ],
            #             "associatedVids": [contact.get('vid')]
            #         },
            #         "properties": [
            #             {
            #                 "value": "{} - Nuevo negocio".format(instance.client.get_full_name()),
            #                 "name": "dealname"
            #             }, {
            #                 "value": "decisionmakerboughtin",  # Deal set in Appointment column
            #                 "name": "dealstage"
            #             },
            #         ]
            #     })
            #
            # else:
            #     return Response(status=status.HTTP_400_BAD_REQUEST, data={"error": _("Deal or company are wrong")})

            if contact:
                instance.hubspot_status = True
                instance.save()
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_400_BAD_REQUEST,
                                data={"error": _("Something went wrong. Contact could not be created")})

    def _get_hubspot_instance(self, lead):
        return Hubspot3(api_key=lead.concessionaire.hubspot_api_key)  # Get concessionarie api key for HBS

    @action(methods=["GET"], detail=False)
    def dashboard_benchmark(self, request, *args, **kwargs):
        useful_filter = request.GET.get('useful', None)
        ids = self.filter_queryset(self.get_queryset())
        ids = [l.id for l in ids]
        qs = LeadMasterPersistente.objects.filter(id__in=ids)
        if useful_filter:
            qs = qs.filter(lead_util=1)

        funnel = qs.aggregate(
            lead_recibido=Sum('lead_recibido'),
            lead_util=Sum('lead_util'),
            ganados=Count('id', filter=Q(resultado='Ganado'))
        )

        origins = qs.values('origen').annotate(
            abiertos=Count('id', filter=Q(resultado__isnull=True)),
            ganados=Count('id', filter=Q(resultado='Ganado')),
            descartados=Count('id', filter=Q(resultado='Descartado')),
            total=Count('id'),
        )

        users = qs.values('asignado').annotate(
            abiertos=Count('id', filter=Q(resultado__isnull=True)),
            ganados=Count('id', filter=Q(resultado='Ganado')),
            descartados=Count('id', filter=Q(resultado='Descartado')),
            total=Count('id'),
        )

        return Response(
            status=status.HTTP_200_OK,
            data={
                'funnel': funnel,
                'origins': origins,
                'users': users,
            }
        )

    @action(methods=["GET"], detail=False)
    def dashboard_service_level(self, request, *args, **kwargs):
        useful_filter = request.GET.get('useful', None)
        ids = self.filter_queryset(self.get_queryset())
        ids = [l.id for l in ids]
        qs = LeadMasterPersistente.objects.filter(id__in=ids)
        if useful_filter:
            qs = qs.filter(lead_util=1)

        kpis = qs.aggregate(
            total=Count('id'),
            tareas_realizadas=Count('id', filter=Q(numero_tareas_realizadas__gt=0)),
            tareas_programadas=Count('id', filter=Q(numero_tareas_programadas__gt=0)),
            sin_tareas_programadas=Count('id', filter=Q(numero_tareas_programadas=0)),
            seguimientos_realizados=Count('id', filter=Q(numero_seguimientos_realizados__gt=0)),
            seguimientos_programados=Count('id', filter=Q(numero_seguimientos_programados__gt=0)),
            sin_seguimientos_programados=Count('id', filter=Q(numero_seguimientos_programados=0)),
            # tareas_atrasadas=Count('id', filter=Q(fecha_ultima_tarea_programada__lt=timezone.now())),
            # seguimientos_atrasados=Count('id', filter=Q(fecha_ultimo_seguimiento_programado__lt=timezone.now()))
        )

        # nuevas_altas = qs.values('created__year', 'created__month', 'created__day').annotate(
        #     total=Count('id')
        # ).order_by('created__year', 'created__month', 'created__day')

        return Response(
            status=status.HTTP_200_OK,
            data={
                'kpis': kpis,
                # 'nuevas_altas': nuevas_altas
            }
        )

    @action(methods=["GET"], detail=False)
    def dashboard_voice_service_level(self, request, *args, **kwargs):
        useful_filter = request.GET.get('useful', None)
        ids = self.filter_queryset(self.get_queryset())
        ids = [l.id for l in ids]
        qs = LeadMasterPersistente.objects.filter(id__in=ids)
        if useful_filter:
            qs = qs.filter(lead_util=1)

        qs = qs.filter(fecha_llamada_entrante__gt='')

        totales = qs.aggregate(
            total=Count('id'),
            no_atendidas=Count('id', filter=~Q(status_call='Atendido'))
        )

        llamadas = qs.annotate(
            hour=Substr('fecha_llamada_entrante', 12, 2)
        ).values('hour').annotate(
            total=Count('id'),
            no_atendidas=Count('id', filter=~Q(status_call='Atendido'))
        ).order_by('hour')

        no_atendidos = qs.filter(~Q(status_call='Atendido')).aggregate(
            con_llamadas=Count('id', filter=Q(numero_llamadas_realizadas__gt=0)),
            sin_llamadas=Count('id', filter=Q(numero_llamadas_realizadas=0))
        )

        asignados = qs.filter(
            ~Q(status_call='Atendido')
        ).values('asignado').annotate(
            con_llamadas=Count('id', filter=Q(numero_llamadas_realizadas__gt=0)),
            sin_llamadas=Count('id', filter=Q(numero_llamadas_realizadas=0))
        )

        asa = qs.filter(
            ~Q(status_call='Atendido')
        ).filter(
            fecha_llamada_entrante__isnull=False,
            fecha_llamada_saliente__isnull=False
        ).annotate(
            llamada_entrante=Cast(
                Concat(
                    Substr('fecha_llamada_entrante', 7, 4),
                    Value('-'),
                    Substr('fecha_llamada_entrante', 4, 2),
                    Value('-'),
                    Substr('fecha_llamada_entrante', 1, 2),
                    Value(' '),
                    Substr('fecha_llamada_entrante', 12, 8),
                ), DateTimeField()
            ),
            llamada_saliente=Cast(
                Concat(
                    Substr('fecha_llamada_saliente', 7, 4),
                    Value('-'),
                    Substr('fecha_llamada_saliente', 4, 2),
                    Value('-'),
                    Substr('fecha_llamada_saliente', 1, 2),
                    Value(' '),
                    Substr('fecha_llamada_saliente', 12, 8),
                ), DateTimeField()
            ),
        ).filter(
            llamada_saliente__gte=F('llamada_entrante')
        ).annotate(
            diff=ExpressionWrapper(F('llamada_saliente') - F('llamada_entrante'), output_field=DurationField())
        ).annotate(
            es_menos_una_hora=Case(
                When(diff__lte=datetime.timedelta(hours=1), then=Value(1)),
                default=Value(0),
                output_field=IntegerField()
            )
        ).aggregate(
            tiempo=Avg('diff'),
            menos_una_hora=Sum('es_menos_una_hora'),
            total=Count('id')
        )

        return Response(
            status=status.HTTP_200_OK,
            data={
                'totales': totales,
                'llamadas': llamadas,
                'no_atendidos': no_atendidos,
                'asignados': asignados,
                'asa': asa,

            }
        )

    @action(methods=["GET"], detail=False)
    def dashboard_digital_service_level(self, request, *args, **kwargs):
        useful_filter = request.GET.get('useful', None)
        ids = self.filter_queryset(self.get_queryset())
        ids = [l.id for l in ids]
        qs = LeadMasterPersistente.objects.filter(id__in=ids)
        if useful_filter:
            qs = qs.filter(lead_util=1)

        qs = qs.filter(~Q(medio__in=['Teléfono', 'Presencial']))

        totales = qs.aggregate(
            total=Count('id'),
            no_atendidas=Count('id', filter=Q(
                numero_llamadas_realizadas=0,
                numero_email_enviados=0,
                numero_whatsapp=0
            ))
        )

        con_telefono = qs.filter(telefono__isnull=False).aggregate(
            con_llamadas=Count('id', filter=Q(numero_llamadas_realizadas__gt=0)),
            sin_llamadas=Count('id', filter=Q(numero_llamadas_realizadas=0))
        )

        asignados = qs.values('asignado').annotate(
            con_llamadas=Count('id', filter=Q(numero_llamadas_realizadas__gt=0)),
            sin_llamadas=Count('id', filter=Q(numero_llamadas_realizadas=0))
        )

        asa = qs.filter(
            ~Q(status_call='Atendido')
        ).filter(
            fecha_llamada_saliente__isnull=False
        ).annotate(
            llamada_saliente=Cast(
                Concat(
                    Substr('fecha_llamada_saliente', 7, 4),
                    Value('-'),
                    Substr('fecha_llamada_saliente', 4, 2),
                    Value('-'),
                    Substr('fecha_llamada_saliente', 1, 2),
                    Value(' '),
                    Substr('fecha_llamada_saliente', 12, 8),
                ), DateTimeField()
            ),
        ).filter(
            llamada_saliente__gte=F('created')
        ).annotate(
            diff=ExpressionWrapper(F('llamada_saliente') - F('created'), output_field=DurationField())
        ).annotate(
            es_menos_una_hora=Case(
                When(diff__lte=datetime.timedelta(hours=1), then=Value(1)),
                default=Value(0),
                output_field=IntegerField()
            )
        ).aggregate(
            tiempo=Avg('diff'),
            menos_una_hora=Sum('es_menos_una_hora'),
            total=Count('id')
        )

        return Response(
            status=status.HTTP_200_OK,
            data={
                'totales': totales,
                'con_telefono': con_telefono,
                'asignados': asignados,
                'asa': asa,
            }
        )


class TaskView(mixins.RetrieveModelMixin, mixins.ListModelMixin, mixins.DestroyModelMixin, mixins.UpdateModelMixin,
               mixins.CreateModelMixin, GenericViewSet):
    serializer_class = TaskSerializer
    queryset = Task.objects.all()
    permission_classes = [Or(IsAdminUser, DjangoModelPermissions, IsAuthenticated)]
    filter_fields = {'is_click2call': ['exact']}

    def get_serializer(self, *args, **kwargs):
        if self.action in ['create']:
            self.serializer_class = TaskCreateSerializer
        elif self.action in ['lead_task_create']:
            self.serializer_class = LeadTaskCreateSerializer
        elif self.action in ['update', 'partial_update']:
            self.serializer_class = CommonTaskSerializer
        return super().get_serializer(*args, **kwargs)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        for i in instance.request_set.all():
            self.check_lead_status(request, i.lead)
        return super().update(request, *args, **kwargs)

    def create(self, request, *args, **kwargs):
        req_ids = request.data.get('request', [])
        if not isinstance(req_ids, list):
            if req_ids:
                req_ids = [req_ids]
            else:
                req_ids = []
        if req_ids:
            for i in Request.objects.filter(id__in=req_ids):
                self.check_lead_status(request, i.lead)
        return super().create(request, *args, **kwargs)

    @action(methods=['POST'], detail=False)
    def lead_task_create(self, request, *args, **kwargs):
        return super().create(request, *args, **kwargs)

    is_traking_task_param = openapi.Parameter('is_traking_task', openapi.IN_QUERY, description="is traking task",
                                              type=openapi.TYPE_BOOLEAN)

    @swagger_auto_schema(method='get', manual_parameters=[is_traking_task_param], responses={})
    @action(methods=['GET'], detail=False)
    def options(self, request, *args, **kwargs):

        from .constant import REQUEST_TASK_TYPES, TASK_TYPES

        try:
            is_traking_task = json.loads(request.GET.get('is_traking_task', 0))
        except:
            is_traking_task = 0

        if is_traking_task:
            return Response(REQUEST_TASK_TYPES)
        return Response(TASK_TYPES)


class GasTypeView(mixins.ListModelMixin, mixins.CreateModelMixin, mixins.DestroyModelMixin, GenericViewSet):
    serializer_class = GasTypeSerializer
    queryset = GasType.objects.all().order_by('id')
    permission_classes = [Or(IsAdminUser, DjangoModelPermissions, IsAuthenticated)]


DATE_RANGE = openapi.Parameter(name='date_range', in_=openapi.IN_QUERY, type=openapi.TYPE_STRING)


class LeadCols(mixins.ListModelMixin, LeadGenericViewSet):
    serializer_class = LeadColumnSerializer
    queryset = Lead.objects.all()
    permission_classes = [Or(IsAdminUser, DjangoModelPermissions, IsAuthenticated)]
    filter_class = LeadColsFilter
    search_fields = (
        'status', 'vehicles__brand_model', 'client__name', 'client__surname', 'client__phone', 'client__desk_phone', 'client__email', 'request__task__description',
        'request__task__realization_date', 'request__task__realization_date', 'user__username',
        'user__first_name', 'user__last_name', 'user__phone', 'user__email', 'concessionaire__name',
        'source__data', 'source__channel__slug', 'source__origin__name', 'client__business_name')

    ordering_fields = ('created', 'lead_task_date')

    @swagger_auto_schema(manual_parameters=[DATE_RANGE])
    def list(self, request, *args, **kwargs):
        if hasattr(request.GET, '_mutable'):
            request.GET._mutable = True
        date_range = request.GET.pop('date_range', '')
        if isinstance(date_range, list):
            date_range = date_range[0]
        if request.GET.get('status', '') in ['new', 'pending']:
            request.GET.pop('status', None)
            self.queryset = self.get_queryset().filter(status__in=['new', 'pending']).order_by('id')

        with_concession = request.GET.get("with_concession", False)
        request.GET.pop("with_concession", None)
        if hasattr(request.GET, '_mutable'):
            request.GET._mutable = False
        query = special_date_ranges(date_range)
        if query:
            self.queryset = self.get_queryset().filter(**query)

        if not with_concession:
            self.queryset = self.get_queryset().filter(user=request.user)
        return super().list(request, *args, **kwargs)


class ACDView(mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.CreateModelMixin, mixins.UpdateModelMixin,
              mixins.DestroyModelMixin, GenericViewSet):
    serializer_class = ACDSerializer
    queryset = ACD.objects.filter(lead__isnull=True).order_by("id")
    permission_classes = [Or(IsAdminUser, DjangoModelPermissions, IsAuthenticated)]

    def get_queryset(self, *args, **kwargs):
        if not (self.request.user.is_superuser or self.request.user.is_staff):
            self.queryset = self.queryset.filter(
                Q(lead__concessionaire__userconcession__user__id=self.request.user.id,
                  lead__concessionaire__userconcession__is_concessionaire_admin=True
                  ) | Q(lead__user__id=self.request.user.id) | Q(lead__isnull=True)) \
                .distinct().order_by('id')
        return super().get_queryset()


class OriginView(mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.CreateModelMixin, mixins.UpdateModelMixin,
                 mixins.DestroyModelMixin, GenericViewSet):
    serializer_class = OriginSerializer
    queryset = Origin.objects.all().order_by("id")
    permission_classes = [IsAuthenticated]
    search_fields = ('name',)
    filter_fields = {'source__concession': ['exact']}
    filterset_fields = ('source__concession',)

    def get_queryset(self):
        return super().get_queryset().distinct()


class LeadFullHistoryView(mixins.ListModelMixin, GenericViewSetAux):
    serializer_class = HistorySerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        queryset = Lead.objects.all()
        lookup_url_kwarg = self.lookup_url_kwarg or self.lookup_field

        assert lookup_url_kwarg in self.kwargs, (
                'Expected view %s to be called with a URL keyword argument '
                'named "%s". Fix your URL conf, or set the `.lookup_field` '
                'attribute on the view correctly.' %
                (self.__class__.__name__, lookup_url_kwarg)
        )

        filter_kwargs = {self.lookup_field: self.kwargs[lookup_url_kwarg]}
        obj = get_object_or_404(queryset, **filter_kwargs)
        self.check_object_permissions(self.request, obj)
        return obj

    def get_models_names(self):
        instance = self.get_object()

        tasks_id = instance.request.task.all().values_list('id', flat=True)
        notes_id = instance.note.all().values_list('id', flat=True)
        tags_id = instance.tags.all().values_list('id', flat=True)
        models_names = [
            {'app_name': 'leads', 'model_name': 'HistoricalClient',
             'lead_field_name': instance.client and [instance.client.id] or None},
            {'app_name': 'leads', 'model_name': 'HistoricalTask',
             'lead_field_name': tasks_id or None},
            {'app_name': 'leads', 'model_name': 'HistoricalRequest',
             'lead_field_name': instance.request and [instance.request.id] or None},
            {'app_name': 'leads', 'model_name': 'HistoricalLead',
             'lead_field_name': [instance.id]},
            {'app_name': 'leads', 'model_name': 'HistoricalNote',
             'lead_field_name': notes_id or None},
            {'app_name': 'leads', 'model_name': 'HistoricalVehicle',
             'lead_field_name': instance.vehicles.all() and instance.vehicles.values_list('id', flat=True) or None},
            {'app_name': 'tags_app', 'model_name': 'HistoricalTag',
             'lead_field_name': tags_id or None}
        ]
        return models_names

    def get_queryset(self, *args, **kwargs):
        if not getattr(self, 'swagger_fake_view', False):
            instance = self.get_object()
            models_names = self.get_models_names()
            data = []
            for model in models_names:
                model_class = apps.get_model(model['app_name'], model_name=model['model_name'])
                queryset = []
                if model['lead_field_name']:
                    created_date_delayed = instance.created - datetime.timedelta(seconds=1)
                    queryset = model_class.objects.filter(id__in=model['lead_field_name'],
                                                          history_date__gte=created_date_delayed,
                                                          history_change_reason__isnull=False)
                [data.append({field_name: getattr(instance, field_name, None)
                              for field_name in self.serializer_class.Meta.fields})
                 for instance in queryset]
            data.sort(key=lambda item: item['history_date'], reverse=True)
            self.queryset = data
        else:
            self.queryset = []
        return super().get_queryset()

    @swagger_auto_schema(auto_schema=None)
    def list(self, request, *args, **kwargs):
        if self.action == 'list':
            raise NotImplementedError("Function not available")
        return super().list(request, *args, **kwargs)

    @action(methods=['GET'], detail=True, suffix='List')
    def get_history(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)


class LeadActionView(mixins.ListModelMixin, mixins.CreateModelMixin, GenericViewSet):
    serializer_class = LeadActionSerializer
    queryset = LeadAction.objects.none()
    filter_class = LeadActionFilter
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.action in ['get_column']:
            """
            query = Max(Case(When(lead__leadaction__date__lte=max_today, then='lead__leadaction__date')))
            queryset = self.queryset.annotate(max_leadaction=query)
            dates = set(queryset.values_list('max_leadaction', flat=True))
            self.queryset = self.queryset.filter(
                date__in=dates, user=self.request.user, lead__status__in=[]
            ).distinct()
            """

            lead_queryset = Lead.objects.filter(status__in=['new', 'pending', 'commercial_management', 'tracing'])
            """
            if self.request.user.is_admin:
                concessionaires = self.request.user.userconcession_set.filter(
                    Q(is_concessionaire_admin=True) | Q(is_business_manager=True)
                ).distinct().values_list('concessionaire', flat=True)
                lead_queryset = lead_queryset.filter(
                    Q(concessionaire__id__in=concessionaires) | Q(user__id=self.request.user.id)
                ).distinct().order_by('id')
            
            else:
            """
            lead_queryset = lead_queryset.filter(user=self.request.user)

            self.queryset = self.queryset.filter(
                id__in=[i.last_lead_action.id for i in lead_queryset if i.last_lead_action]).distinct().order_by('date')

        return super().get_queryset()

    @action(methods=['GET'], detail=False)
    def get_column(self, request, *args, **kwargs):
        """
        Get all lead actions for current user. Show on column for paragraph "My leads"
        """
        return super().list(request, *args, **kwargs)


class LeadCalendarView(GenericViewSet):
    queryset = LeadCalendar.objects.all().order_by("id")
    permission_classes = [Or(IsAdminUser, DjangoModelPermissions, IsAuthenticated)]
    serializer_class = LeadCalendarSerializer
    filter_class = LeadCalendarFilter

    def list(self, request, *args, **kwargs):
        if request.user.is_superuser:
            qs = self.filter_queryset(self.get_queryset())
        else:
            qs = self.filter_queryset(self.get_queryset()).filter(
                Q(lead__concessionaire__userconcession__user__id=self.request.user.id,
                  lead__concessionaire__userconcession__is_concessionaire_admin=True
                  )  | Q(lead__user__id=self.request.user.id)).distinct()
        serializer = self.get_serializer(qs, many=True)
        return Response(status=status.HTTP_200_OK, data=serializer.data)

    @action(methods=['GET'], detail=False)
    def kpis(self, request, *args, **kwargs):
        if request.user.is_superuser:
            qs = self.filter_queryset(self.get_queryset())
        else:
            qs = self.filter_queryset(self.get_queryset()).filter(
                Q(lead__concessionaire__userconcession__user__id=self.request.user.id,
                  lead__concessionaire__userconcession__is_concessionaire_admin=True
                  )  | Q(lead__user__id=self.request.user.id)).distinct()

        not_attended = qs.filter(type='pending_lead').count()
        pending_task = qs.filter(type='pending_task', date__gte=timezone.now()).count()
        delayed_task = qs.filter(type='pending_task', date__lt=timezone.now()).count()
        pending_tracking = qs.filter(type='pending_traking', date__gte=timezone.now()).count()
        delayed_tracking = qs.filter(type='pending_traking', date__lt=timezone.now()).count()

        self.filter_class = TaskFilter
        if request.user.is_superuser:
            tqs = self.filter_queryset(Task.objects.all())
        else:
            tqs = self.filter_queryset(Task.objects.filter(
                Q(lead__concessionaire__userconcession__user__id=self.request.user.id,
                  lead__concessionaire__userconcession__is_concessionaire_admin=True
                  ) | Q(lead__user__id=self.request.user.id)).distinct())

        pending_task_done = tqs.filter(is_traking_task=False,
                                       realization_date__lte=F('planified_realization_date')).count()
        delayed_task_done = tqs.filter(is_traking_task=False,
                                       realization_date__gt=F('planified_realization_date')).count()
        pending_tracking_done = tqs.filter(is_traking_task=True,
                                           realization_date__lte=F('planified_realization_date')).count()
        delayed_tracking_done = tqs.filter(is_traking_task=True,
                                           realization_date__gt=F('planified_realization_date')).count()

        self.filter_class = LeadFilter
        if request.user.is_superuser:
            lqs = self.filter_queryset(Lead.objects.all())
        else:
            lqs = self.filter_queryset(Lead.objects.filter(
                Q(concessionaire__user__id=self.request.user.id,
                  concessionaire__userconcession__is_concessionaire_admin=True
                  )  | Q(user__id=self.request.user.id)).distinct())

        attended = lqs.filter(~Q(status='new')).count()

        return Response(
            status=status.HTTP_200_OK,
            data={
                'kpis': [
                    {
                        'name': _('General'),
                        'kpis': [
                            {
                                'name': _('No atendidos'),
                                'second_name': _('Leads atendidos'),
                                'value': not_attended,
                                'done': attended,
                                'filter_key': 'not_attended',
                                'danger': False
                            },
                            {
                                'name': _('Tareas pendientes'),
                                'second_name': _('Realizadas'),
                                'value': pending_task,
                                'done': pending_task_done,
                                'filter_key': 'pending_task',
                                'danger': False
                            },
                            {
                                'name': _('Tareas atrasadas'),
                                'second_name': _('Realizadas'),
                                'value': delayed_task,
                                'done': delayed_task_done,
                                'filter_key': 'delayed_task',
                                'danger': delayed_task > 0
                            },
                            {
                                'name': _('Seguimientos pendientes'),
                                'second_name': _('Realizadas'),
                                'value': pending_tracking,
                                'done': pending_tracking_done,
                                'filter_key': 'pending_tracking',
                                'danger': False
                            },
                            {
                                'name': _('Seguimientos atrasados'),
                                'second_name': _('Realizadas'),
                                'value': delayed_tracking,
                                'done': delayed_tracking_done,
                                'filter_key': 'delayed_tracking',
                                'danger': delayed_tracking > 0
                            },

                        ]
                    }
                ]
            }
        )


@method_decorator(transaction.non_atomic_requests, name='dispatch')
class LeadImporterView(GenericViewSet):
    authentication_classes = [authentication.TokenAuthentication]
    permission_classes = [IsAdminUser]
    queryset = Lead.objects.all().order_by("id")
    serializer_class = LeadImporterSerializer

    @action(detail=False, methods=['get', ])
    def download_template(self, request):

        from django.http import HttpResponse

        output = io.BytesIO()
        filename = 'Template_Leads.xlsx'

        # Preparamos la Hoja de cálculo para rellenarla
        wb = xlsxwriter.Workbook(output)
        ws = wb.add_worksheet('Leads')

        create_excel_template(self.get_serializer(), ws, wb)

        wb.close()
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response

    @action(detail=False, methods=['post', ])
    def upload(self, request):

        filename = 'Result_Leads.xlsx'

        SUCCESS_TOKEN = "OK"
        ERROR_TOKEN = "FAIL"

        validations = []
        columns_names = []

        # Obtenemos el archivo
        if request.method == 'POST' and request.FILES['file']:

            file = request.FILES['file']
            data = file.read()

            output = io.BytesIO()
            wb = xlsxwriter.Workbook(output)
            output_ws = wb.add_worksheet('Leads Imported Result')

            columns = [k for k, v in self.get_serializer().get_fields().items() if not v.write_only]
            imported_lead_id_colum_key = len(columns)
            imported_result_colum_key = len(columns) + 1
            imported_error_colum_key = len(columns) + 2

            create_excel_template(self.get_serializer(), output_ws, wb)

            try:
                workbook = xlrd.open_workbook(file_contents=data)
                worksheet = workbook.sheet_by_index(0)
            except Exception as e:
                raise ValidationError({"detail": _("Error en la apertura de xlsx, ")})

            for col in range(worksheet.ncols):
                columns_names.append(worksheet.cell_value(0, col))

            # Importación masiva
            for row in range(1, worksheet.nrows):
                data = {}
                for col in range(worksheet.ncols):

                    # Obtenemos el valor
                    cell_value = worksheet.cell_value(row, col)

                    if isinstance(cell_value, float):
                        try:
                            cell_value = int(cell_value)
                        except ValueError:
                            cell_value = format(cell_value, '.0f')

                    data[columns_names[col]] = cell_value
                    add_value_to_excel(cell_value, output_ws, row, col)

                # Sobreescribir si hace falta tratar los datos obtenidos
                data = {k: v for k, v in data.items() if v not in (None, "")}

                try:

                    if data.get('imported_result') != SUCCESS_TOKEN:

                        # Damos formato a los datos
                        data = self.get_serializer(data=data).validate(data)
                        serializer = self.get_serializer(data=data)
                        serializer.is_valid(raise_exception=True)

                        try:
                            serializer.save()
                        except Exception as error:

                            add_value_to_excel("-", output_ws, row, imported_lead_id_colum_key)
                            add_value_to_excel(str(error), output_ws, row, imported_error_colum_key)
                            add_value_to_excel(ERROR_TOKEN, output_ws, row, imported_result_colum_key)
                            continue

                        # Obtenemos los ids
                        lead_id = serializer.instance.pk
                        add_value_to_excel(lead_id, output_ws, row, imported_lead_id_colum_key)
                        add_value_to_excel("-", output_ws, row, imported_error_colum_key)
                        add_value_to_excel(SUCCESS_TOKEN, output_ws, row, imported_result_colum_key)

                except ValidationError as errors:
                    validations = []
                    for err in errors.get_full_details():
                        for field in errors.get_full_details()[err]:
                            validations.append({
                                "field": err,
                                'message': field["message"],
                            })

                    add_value_to_excel("-", output_ws, row, imported_lead_id_colum_key)
                    add_value_to_excel(json.dumps(validations), output_ws, row, imported_error_colum_key)
                    add_value_to_excel(ERROR_TOKEN, output_ws, row, imported_result_colum_key)

            wb.close()
            output.seek(0)

            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=%s' % filename
            return response

        raise ValidationError({"detail": _("Debe añadir un fichero con formato xlsx, ")})


class LeadWhastAppMessageView(mixins.CreateModelMixin,
                              mixins.UpdateModelMixin,
                              mixins.ListModelMixin,
                              mixins.RetrieveModelMixin,
                              mixins.DestroyModelMixin, GenericViewSet):
    serializer_class = WhatsAppSerializer
    queryset = LeadWhatsAppMessage.objects.all()
    permission_classes = (IsAuthenticated,)
